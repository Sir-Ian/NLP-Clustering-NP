import pandas as pd
import re
from openpyxl import load_workbook
import os

# Define the directory path
dir_path = "/Users/peretzcik/Library/Mobile Documents/com~apple~CloudDocs/Downloads/LiveChatData"
os.chdir(dir_path)

# Define the file name
file_name = 'TestData.xlsx'

# Load the workbook
book = load_workbook(file_name)

# Load the worksheets
data_ws = book['Sheet1']
output_ws = book['Sheet2']

# Define the patterns for the different sections
patterns = [
    ("Concern", r"Concern (\d+): ([^\n]*)"),
    ("Response", r"Response (\d+): ([^\n]*)"),
    ("Topic", r"Topic (\d+): ([^\n]*)"),
    ("Subtopic", r"Subtopic (\d+): ([^\n]*)"),
    ("Keywords", r"Keywords (\d+): ([^\n]*)"),
    ("Sentiment", r"Sentiment (\d+): ([^\n]*)"),
    ("Resolution", r"Resolution (\d+): ([^\n]*)"),
    ("Improvement Area", r"Improvement Area (\d+): ([^\n]*)"),
    ("Barrier", r"Barrier: ([^\n]*)"),
    ("Tone", r"Tone: ([^\n]*)"),
    ("Status", r"Status: ([^\n]*)"),
    ("Product Category", r"Product Category: ([^\n]*)"),
    ("Specific Product", r"Specific Product: ([^\n]*)"),
]

def parse_entry(entry):
    # Initialize a dict to hold the parsed data
    data = {name: [] for name, _ in patterns}

    # Apply the patterns
    for name, pattern in patterns:
        matches = re.findall(pattern, entry, re.MULTILINE)
        if matches:
            # Handle multi-line data (multiple concerns, etc.)
            if isinstance(matches[0], tuple):
                for match in matches:
                    idx, value = match
                    data[name].append(value)
            else:
                value = matches[0]
                data[name].append(value)
        else:
            data[name].append(None)

    # Find the longest list in the data
    max_length = max(len(lst) for lst in data.values())

    # List of non-indexed fields
    non_indexed_fields = ["Barrier", "Tone", "Status", "Product Category", "Specific Product"]

    # Extend all lists to be of the max length
    for name, lst in data.items():
        # If field is non-indexed, repeat the value for all concerns
        if name in non_indexed_fields and lst != [None]:
            data[name] = lst * max_length
        # If field is indexed or None, fill the list with None
        else:
            lst += [None] * (max_length - len(lst))

    return pd.DataFrame(data)


# Add a header for the chat_id on Sheet2
output_ws.cell(row=1, column=1, value="chat_id")

# Process each entry in column B (2nd column, 1-indexed)
for row in range(2, data_ws.max_row + 1):
    chat_id = data_ws.cell(row=row, column=1).value  # Read the chat_id from column A
    entry = data_ws.cell(row=row, column=2).value  # Read the entry from column B
    result = parse_entry(entry)

    # Write the result back to 'Sheet2'
    for i, record in result.iterrows():
        output_row = output_ws.max_row + 1
        output_ws.cell(row=output_row, column=1, value=chat_id)  # Write the chat_id to column A
        for j, (column_name, value) in enumerate(record.items(), start=1):
            output_ws.cell(row=output_row, column=j + 1, value=value)  # Adjust column indexing to start from column B

# Save the workbook
book.save(file_name)

print("Task completed successfully!")
