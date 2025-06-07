import openai
import time
from openpyxl import load_workbook
import os
import shutil
from datetime import datetime

# Initialize the OpenAI API
openai.api_key = 'PUT_KEY_HERE'

dir_path = "/Users/peretzcik/Library/Mobile Documents/com~apple~CloudDocs/Downloads/LiveChatData"
os.chdir(dir_path)

# Backup the Excel file
shutil.copy2('LiveChatData.xlsx', 'LiveChatData_backup.xlsx')

# Load the workbook
book = load_workbook('LiveChatData.xlsx')

# Load the worksheets
post_chat_data_ws = book['Post Chat Data']
prompt_ws = book['Prompt']

# Get the prompt prefix
prompt_prefix = prompt_ws['B1'].value

# Counter for rows processed
row_counter = 0

# Process each conversation in column H (8th column, 1-indexed)
for row in range(2, post_chat_data_ws.max_row + 1):
    if not post_chat_data_ws.cell(row=row, column=16).value:  # Column P is the 16th column
        try:
            # Concatenate prompt prefix with conversation
            prompt = prompt_prefix + str(post_chat_data_ws.cell(row=row, column=8).value)  # Column H is the 8th column

            # Get the response from OpenAI API
            response = openai.ChatCompletion.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "This is your assistant speaking."},
                    {"role": "user", "content": prompt}
                ]
            )

            # Store the response in column P
            post_chat_data_ws.cell(row=row, column=16, value=response['choices'][0]['message']['content'])

            # Increment row counter
            row_counter += 1

            if row_counter == 50:
                # Save after each 50 successful API calls and reset the counter
                book.save('LiveChatData.xlsx')
                row_counter = 0

        except Exception as e:
            print(f"Encountered an error at row {row}: {e}")
            post_chat_data_ws.cell(row=row, column=16, value="error")

        # Decrement the counter in R1
        post_chat_data_ws['R1'] = post_chat_data_ws['R1'].value - 1

        # Generate timestamp for each iteration
        current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

        print(f'Completed row {row} at {current_timestamp}. Remaining iterations: {post_chat_data_ws["R1"].value}')

        # Check the counter
        if post_chat_data_ws['R1'].value <= 0:
            break

    # Sleep for a second to avoid hitting API rate limit
    time.sleep(1)

# Save the workbook one final time after the loop finishes
book.save('LiveChatData.xlsx')

print("Script completed successfully.")
